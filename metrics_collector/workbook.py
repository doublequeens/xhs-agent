from collections import Counter
from datetime import datetime
import math
from pathlib import Path
import re
from typing import Any, Callable, TypeVar
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from metrics_collector.matcher import normalize_title
from metrics_collector.models import ExportedMetrics


class WorkbookFormatError(ValueError):
    pass


HEADERS = (
    "笔记标题",
    "首次发布时间",
    "体裁",
    "曝光",
    "观看量",
    "封面点击率",
    "点赞",
    "评论",
    "收藏",
    "涨粉",
    "分享",
    "人均观看时长",
    "弹幕",
)

_COUNT_FIELDS = {
    "曝光": "impressions",
    "观看量": "views",
    "点赞": "likes",
    "评论": "comments",
    "收藏": "saves",
    "涨粉": "followers_gained",
    "分享": "shares",
    "弹幕": "danmaku_count",
}
_INTEGER_STRING = re.compile(r"(?:\d+|\d{1,3}(?:,\d{3})+)")
_DURATION_STRING = re.compile(r"(\d+)\s*(?:s|秒)")
_PUBLICATION_FORMATS = (
    "%Y年%m月%d日%H时%M分%S秒",
    "%Y年%m月%d日%H时%M分",
    "%Y-%m-%d %H:%M",
)
_WORKBOOK_READ_ERRORS = (
    BadZipFile,
    EOFError,
    InvalidFileException,
    KeyError,
    OSError,
    ParseError,
    ValueError,
)
_T = TypeVar("_T")


def _is_unavailable(value: Any) -> bool:
    return value is None or (
        isinstance(value, str) and value.strip() in {"", "-"}
    )


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in row
    )


def _error(row_number: int, field: str, detail: str) -> WorkbookFormatError:
    return WorkbookFormatError(
        f"row {row_number} field {field}: {detail}"
    )


def _workbook_error(path: Path, detail: str) -> WorkbookFormatError:
    return WorkbookFormatError(f"workbook {path}: {detail}")


def _header_candidate(
    row: tuple[Any, ...],
    row_number: int,
) -> tuple[dict[str, int] | None, Counter]:
    required = set(HEADERS)
    counts = Counter(value for value in row if value in required)
    if len(counts) != len(HEADERS):
        return None, counts
    duplicates = [header for header in HEADERS if counts[header] > 1]
    if duplicates:
        raise _error(
            row_number,
            duplicates[0],
            f"duplicate required header: {duplicates[0]}",
        )
    return {
        header: next(
            index for index, value in enumerate(row) if value == header
        )
        for header in HEADERS
    }, counts


def _raise_missing_header(
    path: Path,
    best_partial: tuple[int, int, Counter] | None,
) -> None:
    counts = best_partial[2] if best_partial is not None else Counter()
    if best_partial is not None:
        row_number = best_partial[1]
        duplicates = [
            header for header in HEADERS if counts.get(header, 0) > 1
        ]
        if duplicates:
            raise _error(
                row_number,
                duplicates[0],
                f"duplicate required header: {duplicates[0]}",
            )
    missing = [header for header in HEADERS if header not in counts]
    raise _workbook_error(
        path,
        f"missing required headers: {', '.join(missing)}",
    )


def _updated_best_partial(
    current: tuple[int, int, Counter] | None,
    row_number: int,
    counts: Counter,
) -> tuple[int, int, Counter]:
    candidate = (len(counts), row_number, counts)
    if current is None or candidate[0] > current[0]:
        return candidate
    return current


def _run_workbook_pass(
    path: Path,
    *,
    data_only: bool,
    operation: Callable[[Any], _T],
) -> _T:
    workbook = None
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=data_only,
        )
        if workbook.active is None:
            raise _workbook_error(path, "missing active worksheet")
        return operation(workbook.active)
    except WorkbookFormatError:
        raise
    except _WORKBOOK_READ_ERRORS as exc:
        raise _workbook_error(
            path,
            f"could not read workbook: {exc}",
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()


def _validate_formula_free_workbook(
    path: Path,
) -> tuple[int, dict[str, int]]:
    def validate(sheet: Any) -> tuple[int, dict[str, int]]:
        header: tuple[int, dict[str, int]] | None = None
        best_partial: tuple[int, int, Counter] | None = None
        for row_number, cells in enumerate(
            sheet.iter_rows(values_only=False),
            start=1,
        ):
            values = tuple(cell.value for cell in cells)
            columns, counts = _header_candidate(values, row_number)
            if columns is not None:
                if header is not None:
                    raise _error(
                        row_number,
                        "headers",
                        "multiple header rows found at "
                        f"rows {header[0]} and {row_number}",
                    )
                header = (row_number, columns)
                continue
            if header is None:
                best_partial = _updated_best_partial(
                    best_partial,
                    row_number,
                    counts,
                )
                continue
            for field, column_index in header[1].items():
                if (
                    column_index < len(cells)
                    and cells[column_index].data_type == "f"
                ):
                    raise _error(
                        row_number,
                        field,
                        "formula cells are not allowed",
                    )
        if header is None:
            _raise_missing_header(path, best_partial)
        return header

    return _run_workbook_pass(
        path,
        data_only=False,
        operation=validate,
    )


def _parse_count(value: Any, row_number: int, field: str) -> int | None:
    if _is_unavailable(value):
        return None
    if isinstance(value, bool):
        raise _error(row_number, field, "expected a nonnegative integer")
    if isinstance(value, int):
        parsed = value
    elif isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            raise _error(row_number, field, "expected a nonnegative integer")
        parsed = int(value)
    elif isinstance(value, str):
        text = value.strip()
        if _INTEGER_STRING.fullmatch(text) is None:
            raise _error(row_number, field, "expected a nonnegative integer")
        parsed = int(text.replace(",", ""))
    else:
        raise _error(row_number, field, "expected a nonnegative integer")
    if parsed < 0:
        raise _error(row_number, field, "expected a nonnegative integer")
    return parsed


def _parse_percent(
    value: Any,
    row_number: int,
    field: str,
) -> float | None:
    if _is_unavailable(value):
        return None
    if isinstance(value, bool):
        raise _error(row_number, field, "expected a percentage from 0 to 1")

    try:
        if isinstance(value, (int, float)):
            parsed = float(value)
        elif isinstance(value, str):
            text = value.strip()
            if text.endswith("%"):
                parsed = float(text[:-1].strip()) / 100
            else:
                parsed = float(text)
        else:
            raise TypeError
    except (TypeError, ValueError):
        raise _error(
            row_number,
            field,
            "expected a percentage from 0 to 1",
        ) from None

    if not math.isfinite(parsed) or not 0 <= parsed <= 1:
        raise _error(row_number, field, "expected a percentage from 0 to 1")
    return parsed


def _parse_publication_time(
    value: Any,
    supplied_timezone: ZoneInfo,
    row_number: int,
    field: str,
) -> datetime:
    if _is_unavailable(value):
        raise _error(row_number, field, "publication time is required")
    if isinstance(value, bool):
        raise _error(row_number, field, "invalid publication time")

    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str):
        text = value.strip()
        parsed = None
        for date_format in _PUBLICATION_FORMATS:
            try:
                parsed = datetime.strptime(text, date_format)
                break
            except ValueError:
                continue
        if parsed is None:
            raise _error(row_number, field, "invalid publication time")
    else:
        raise _error(row_number, field, "invalid publication time")

    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=supplied_timezone)
    return parsed.astimezone(supplied_timezone)


def _parse_duration(value: Any, row_number: int, field: str) -> int | None:
    if _is_unavailable(value):
        return None
    if isinstance(value, bool):
        raise _error(row_number, field, "expected nonnegative seconds")
    if isinstance(value, int):
        parsed = value
    elif isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            raise _error(row_number, field, "expected nonnegative seconds")
        parsed = int(value)
    elif isinstance(value, str):
        match = _DURATION_STRING.fullmatch(value.strip())
        if match is None:
            raise _error(row_number, field, "expected nonnegative seconds")
        parsed = int(match.group(1))
    else:
        raise _error(row_number, field, "expected nonnegative seconds")
    if parsed < 0:
        raise _error(row_number, field, "expected nonnegative seconds")
    return parsed


def _cell(
    row: tuple[Any, ...],
    columns: dict[str, int],
    field: str,
) -> Any:
    index = columns[field]
    return row[index] if index < len(row) else None


def _parse_row(
    row: tuple[Any, ...],
    row_number: int,
    columns: dict[str, int],
    supplied_timezone: ZoneInfo,
) -> ExportedMetrics:
    title_value = _cell(row, columns, "笔记标题")
    if not isinstance(title_value, str) or not title_value.strip():
        raise _error(row_number, "笔记标题", "title is required")
    title = title_value.strip()
    published_at = _parse_publication_time(
        _cell(row, columns, "首次发布时间"),
        supplied_timezone,
        row_number,
        "首次发布时间",
    )

    parsed_counts = {
        attribute: _parse_count(
            _cell(row, columns, field),
            row_number,
            field,
        )
        for field, attribute in _COUNT_FIELDS.items()
    }
    return ExportedMetrics(
        title=title,
        published_at=published_at,
        cover_click_rate=_parse_percent(
            _cell(row, columns, "封面点击率"),
            row_number,
            "封面点击率",
        ),
        avg_watch_time_seconds=_parse_duration(
            _cell(row, columns, "人均观看时长"),
            row_number,
            "人均观看时长",
        ),
        **parsed_counts,
    )


def parse_metrics_workbook(
    path: Path,
    timezone: ZoneInfo,
) -> list[ExportedMetrics]:
    expected_header = _validate_formula_free_workbook(path)

    def parse(sheet: Any) -> list[ExportedMetrics]:
        header: tuple[int, dict[str, int]] | None = None
        best_partial: tuple[int, int, Counter] | None = None
        parsed_rows = []
        identities: dict[tuple[str, datetime], int] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(values_only=True),
            start=1,
        ):
            columns, counts = _header_candidate(row, row_number)
            if columns is not None:
                if header is not None:
                    raise _error(
                        row_number,
                        "headers",
                        "multiple header rows found at "
                        f"rows {header[0]} and {row_number}",
                    )
                header = (row_number, columns)
                if header != expected_header:
                    raise _workbook_error(
                        path,
                        "header region changed between validation and value "
                        "passes",
                    )
                continue
            if header is None:
                best_partial = _updated_best_partial(
                    best_partial,
                    row_number,
                    counts,
                )
                continue
            if _is_empty_row(row):
                continue
            parsed = _parse_row(row, row_number, header[1], timezone)
            identity = (normalize_title(parsed.title), parsed.published_at)
            if identity in identities:
                original_row = identities[identity]
                raise _error(
                    row_number,
                    "笔记标题",
                    "duplicate title and publication time; "
                    f"original row {original_row}, duplicate row {row_number}",
                )
            identities[identity] = row_number
            parsed_rows.append(parsed)
        if header is None:
            _raise_missing_header(path, best_partial)
        return parsed_rows

    return _run_workbook_pass(
        path,
        data_only=True,
        operation=parse,
    )
