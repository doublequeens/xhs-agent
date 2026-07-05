from datetime import datetime, timedelta, timezone
from math import inf, isfinite, nan
from pathlib import Path
from zipfile import ZipFile
from zoneinfo import ZoneInfo

from openpyxl import Workbook
import pytest

from metrics_collector.models import ExportedMetrics
from metrics_collector.workbook import (
    WorkbookFormatError,
    parse_metrics_workbook,
)


TZ = ZoneInfo("Asia/Shanghai")
HEADERS = [
    "笔记标题",
    "首次发布时间",
    "体裁",
    "曝光",
    "观看量",
    "封面点击率",
    "点赞",
    "评论",
    "收藏",
    "涨粉",
    "分享",
    "人均观看时长",
    "弹幕",
]
VALID_ROW = [
    "室外补防晒技巧",
    "2026年05月16日09时55分23秒",
    "图文",
    1191,
    72,
    0.06,
    10,
    2,
    3,
    4,
    5,
    17,
    6,
]


def build_workbook(
    tmp_path,
    rows,
    *,
    headers=HEADERS,
    leading_rows=None,
    filename="metrics.xlsx",
):
    path = tmp_path / filename
    workbook = Workbook()
    sheet = workbook.active
    for row in leading_rows or []:
        sheet.append(row)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def replace(row, header, value):
    changed = list(row)
    changed[HEADERS.index(header)] = value
    return changed


def test_parse_official_workbook_finds_header_after_notice_and_maps_types(
    tmp_path,
):
    path = build_workbook(
        tmp_path,
        [VALID_ROW],
        leading_rows=[
            ["数据说明：以下指标每日更新"],
            ["请以首次发布时间为准"],
        ],
    )

    rows = parse_metrics_workbook(path, TZ)

    assert rows == [
        ExportedMetrics(
            title="室外补防晒技巧",
            published_at=datetime(2026, 5, 16, 9, 55, 23, tzinfo=TZ),
            impressions=1191,
            views=72,
            cover_click_rate=0.06,
            likes=10,
            comments=2,
            saves=3,
            followers_gained=4,
            shares=5,
            avg_watch_time_seconds=17,
            danmaku_count=6,
        )
    ]


def test_maps_required_columns_by_name_and_ignores_extra_columns(tmp_path):
    headers = ["无关列", *reversed(HEADERS), "另一个无关列"]
    values = dict(zip(HEADERS, VALID_ROW, strict=True))
    row = ["ignore", *(values[header] for header in reversed(HEADERS)), 999]
    path = build_workbook(tmp_path, [row], headers=headers)

    parsed = parse_metrics_workbook(path, TZ)[0]

    assert parsed.title == VALID_ROW[0]
    assert parsed.impressions == 1191
    assert parsed.danmaku_count == 6


def test_dash_blank_and_none_are_none_but_numeric_and_string_zero_remain_zero(
    tmp_path,
):
    row = [
        "新笔记",
        "2026年07月05日13时29分55秒",
        "图文",
        "-",
        " ",
        None,
        0,
        "0",
        0.0,
        " - ",
        None,
        "0秒",
        0,
    ]
    path = build_workbook(tmp_path, [row])

    parsed = parse_metrics_workbook(path, TZ)[0]

    assert parsed.impressions is None
    assert parsed.views is None
    assert parsed.cover_click_rate is None
    assert parsed.likes == 0
    assert parsed.comments == 0
    assert parsed.saves == 0
    assert parsed.followers_gained is None
    assert parsed.shares is None
    assert parsed.avg_watch_time_seconds == 0
    assert parsed.danmaku_count == 0


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (1234, 1234),
        (1234.0, 1234),
        ("1,234", 1234),
        (" 1,234 ", 1234),
        ("0", 0),
    ],
)
def test_count_fields_accept_supported_integer_forms(tmp_path, value, expected):
    path = build_workbook(
        tmp_path,
        [replace(VALID_ROW, "曝光", value)],
        filename=f"count-{expected}.xlsx",
    )

    assert parse_metrics_workbook(path, TZ)[0].impressions == expected


@pytest.mark.parametrize(
    "value",
    [True, False, -1, -1.0, 1.5, "1.5", "1,23", "abc", nan, inf, -inf],
)
def test_count_fields_reject_invalid_values_with_row_and_field(
    tmp_path,
    monkeypatch,
    value,
):
    row = replace(VALID_ROW, "曝光", value)
    if isinstance(value, float) and not isfinite(value):
        workbook = FakeWorkbook([HEADERS, row])
        monkeypatch.setattr(
            "metrics_collector.workbook.load_workbook",
            lambda *args, **kwargs: workbook,
        )
        path = tmp_path / "not-read.xlsx"
    else:
        path = build_workbook(
            tmp_path,
            [row],
            filename=f"invalid-count-{repr(value)}.xlsx",
        )

    with pytest.raises(
        WorkbookFormatError,
        match=r"row 2.*曝光",
    ):
        parse_metrics_workbook(path, TZ)


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (0.06, 0.06),
        (0, 0.0),
        (0.0, 0.0),
        (1, 1.0),
        ("6%", 0.06),
        (" 6 % ", 0.06),
        ("0.06", 0.06),
        ("0", 0.0),
    ],
)
def test_percent_accepts_supported_variants(tmp_path, value, expected):
    path = build_workbook(
        tmp_path,
        [replace(VALID_ROW, "封面点击率", value)],
        filename=f"percent-{repr(value)}.xlsx",
    )

    assert parse_metrics_workbook(path, TZ)[0].cover_click_rate == pytest.approx(
        expected
    )


@pytest.mark.parametrize(
    "value",
    [True, -0.01, 1.01, "101%", "-1%", "nan", "inf", nan, inf, -inf, "abc"],
)
def test_percent_rejects_out_of_range_nonfinite_and_malformed_values(
    tmp_path,
    monkeypatch,
    value,
):
    row = replace(VALID_ROW, "封面点击率", value)
    if isinstance(value, float) and not isfinite(value):
        workbook = FakeWorkbook([HEADERS, row])
        monkeypatch.setattr(
            "metrics_collector.workbook.load_workbook",
            lambda *args, **kwargs: workbook,
        )
        path = tmp_path / "not-read.xlsx"
    else:
        path = build_workbook(
            tmp_path,
            [row],
            filename=f"invalid-percent-{repr(value)}.xlsx",
        )

    with pytest.raises(
        WorkbookFormatError,
        match=r"row 2.*封面点击率",
    ):
        parse_metrics_workbook(path, TZ)


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (
            "2026年05月16日09时55分23秒",
            datetime(2026, 5, 16, 9, 55, 23, tzinfo=TZ),
        ),
        (
            "2026年05月16日09时55分",
            datetime(2026, 5, 16, 9, 55, tzinfo=TZ),
        ),
        (
            "2026-05-16 09:55",
            datetime(2026, 5, 16, 9, 55, tzinfo=TZ),
        ),
        (
            datetime(2026, 5, 16, 9, 55, 23),
            datetime(2026, 5, 16, 9, 55, 23, tzinfo=TZ),
        ),
    ],
)
def test_publication_time_accepts_supported_variants(tmp_path, value, expected):
    path = build_workbook(
        tmp_path,
        [replace(VALID_ROW, "首次发布时间", value)],
        filename=f"time-{expected:%H%M%S}.xlsx",
    )

    assert parse_metrics_workbook(path, TZ)[0].published_at == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (17, 17),
        (17.0, 17),
        ("17s", 17),
        (" 17 s ", 17),
        ("17秒", 17),
        (" 17 秒 ", 17),
        ("0s", 0),
    ],
)
def test_watch_duration_accepts_supported_variants(tmp_path, value, expected):
    path = build_workbook(
        tmp_path,
        [replace(VALID_ROW, "人均观看时长", value)],
        filename=f"duration-{repr(value)}.xlsx",
    )

    assert (
        parse_metrics_workbook(path, TZ)[0].avg_watch_time_seconds == expected
    )


@pytest.mark.parametrize(
    "value",
    [True, -1, -1.0, 1.5, "1.5s", "-1秒", "17", "seconds", nan, inf],
)
def test_watch_duration_rejects_invalid_values(tmp_path, monkeypatch, value):
    row = replace(VALID_ROW, "人均观看时长", value)
    if isinstance(value, float) and not isfinite(value):
        workbook = FakeWorkbook([HEADERS, row])
        monkeypatch.setattr(
            "metrics_collector.workbook.load_workbook",
            lambda *args, **kwargs: workbook,
        )
        path = tmp_path / "not-read.xlsx"
    else:
        path = build_workbook(
            tmp_path,
            [row],
            filename=f"invalid-duration-{repr(value)}.xlsx",
        )

    with pytest.raises(
        WorkbookFormatError,
        match=r"row 2.*人均观看时长",
    ):
        parse_metrics_workbook(path, TZ)


@pytest.mark.parametrize(
    ("headers", "message"),
    [
        (
            [header for header in HEADERS if header != "曝光"],
            "missing required headers.*曝光",
        ),
        (
            ["标题已改名" if header == "笔记标题" else header for header in HEADERS],
            "missing required headers.*笔记标题",
        ),
        (
            [*HEADERS, "曝光"],
            "duplicate required header.*曝光",
        ),
    ],
)
def test_missing_changed_or_duplicate_required_headers_reject_workbook(
    tmp_path,
    headers,
    message,
):
    path = build_workbook(tmp_path, [], headers=headers)

    with pytest.raises(WorkbookFormatError, match=message):
        parse_metrics_workbook(path, TZ)


def test_multiple_header_rows_are_ambiguous(tmp_path):
    path = build_workbook(tmp_path, [HEADERS, VALID_ROW])

    with pytest.raises(WorkbookFormatError, match="multiple header rows"):
        parse_metrics_workbook(path, TZ)


def test_empty_workbook_missing_header_is_a_workbook_level_error(tmp_path):
    path = tmp_path / "empty.xlsx"
    workbook = Workbook()
    workbook.save(path)
    workbook.close()

    with pytest.raises(WorkbookFormatError) as exc_info:
        parse_metrics_workbook(path, TZ)

    message = str(exc_info.value)
    assert str(path) in message
    assert "missing required headers" in message
    assert "row 0" not in message


def test_missing_header_is_a_workbook_level_error(tmp_path):
    path = build_workbook(
        tmp_path,
        [],
        headers=["错误标题", *HEADERS[1:]],
    )

    with pytest.raises(WorkbookFormatError) as exc_info:
        parse_metrics_workbook(path, TZ)

    message = str(exc_info.value)
    assert str(path) in message
    assert "missing required headers" in message
    assert not message.startswith("row ")


def test_skips_wholly_empty_rows_and_parses_every_nonempty_row(tmp_path):
    second = replace(VALID_ROW, "笔记标题", "第二篇")
    second = replace(second, "首次发布时间", "2026-05-17 10:00")
    path = build_workbook(
        tmp_path,
        [[], [None] * len(HEADERS), VALID_ROW, [" "] * len(HEADERS), second],
    )

    rows = parse_metrics_workbook(path, TZ)

    assert [row.title for row in rows] == ["室外补防晒技巧", "第二篇"]


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("笔记标题", None),
        ("笔记标题", "   "),
        ("首次发布时间", None),
        ("首次发布时间", " "),
    ],
)
def test_nonempty_row_missing_identity_rejects_entire_workbook(
    tmp_path,
    field,
    value,
):
    first = replace(VALID_ROW, "笔记标题", "valid first row")
    invalid = replace(VALID_ROW, field, value)
    path = build_workbook(tmp_path, [first, invalid])

    with pytest.raises(
        WorkbookFormatError,
        match=rf"row 3.*{field}",
    ):
        parse_metrics_workbook(path, TZ)


def test_duplicate_normalized_title_and_exact_publication_time_rejects(
    tmp_path,
):
    duplicate = replace(VALID_ROW, "笔记标题", "  室外补防晒技巧  ")
    path = build_workbook(tmp_path, [VALID_ROW, duplicate])

    with pytest.raises(
        WorkbookFormatError,
        match=r"row 3.*笔记标题.*duplicate.*row 2.*row 3",
    ):
        parse_metrics_workbook(path, TZ)


def test_same_title_at_different_publication_time_is_allowed(tmp_path):
    second = replace(VALID_ROW, "首次发布时间", "2026-05-16 09:56")
    path = build_workbook(tmp_path, [VALID_ROW, second])

    assert len(parse_metrics_workbook(path, TZ)) == 2


class FakeWorkbook:
    def __init__(self, rows, *, iterator_type=iter):
        self.active = self
        self.rows = rows
        self.closed = False
        self.iterator_type = iterator_type

    def iter_rows(self, values_only):
        rows = self.rows
        if not values_only:
            rows = [
                tuple(FakeCell(value) for value in row)
                for row in self.rows
            ]
        return self.iterator_type(rows)

    def close(self):
        self.closed = True


class FakeCell:
    def __init__(self, value):
        self.value = value
        self.data_type = (
            "f"
            if isinstance(value, str) and value.startswith("=")
            else "n"
        )


class NoMaterializationIterator:
    def __init__(self, rows):
        self._rows = iter(rows)

    def __iter__(self):
        return self

    def __next__(self):
        return next(self._rows)

    def __length_hint__(self):
        raise AssertionError("row iterator was materialized")


def test_rows_are_consumed_incrementally_without_raw_sheet_materialization(
    monkeypatch,
):
    workbook = FakeWorkbook(
        [HEADERS, VALID_ROW],
        iterator_type=NoMaterializationIterator,
    )
    monkeypatch.setattr(
        "metrics_collector.workbook.load_workbook",
        lambda *args, **kwargs: workbook,
    )

    rows = parse_metrics_workbook(Path("ignored.xlsx"), TZ)

    assert len(rows) == 1


def test_aware_datetime_is_converted_and_workbook_closes_on_success(
    monkeypatch,
):
    aware_time = datetime(
        2026,
        5,
        16,
        1,
        55,
        tzinfo=timezone(timedelta(hours=-7)),
    )
    workbook = FakeWorkbook(
        [HEADERS, replace(VALID_ROW, "首次发布时间", aware_time)]
    )
    monkeypatch.setattr(
        "metrics_collector.workbook.load_workbook",
        lambda *args, **kwargs: workbook,
    )

    parsed = parse_metrics_workbook("ignored.xlsx", TZ)

    assert parsed[0].published_at == datetime(
        2026,
        5,
        16,
        16,
        55,
        tzinfo=TZ,
    )
    assert workbook.closed is True


def test_validation_and_value_workbooks_are_loaded_in_order_and_closed(
    monkeypatch,
):
    workbooks = [
        FakeWorkbook([HEADERS, VALID_ROW]),
        FakeWorkbook([HEADERS, VALID_ROW]),
    ]
    calls = []

    def load(path, *, read_only, data_only):
        calls.append((path, read_only, data_only))
        return workbooks[len(calls) - 1]

    monkeypatch.setattr("metrics_collector.workbook.load_workbook", load)

    parse_metrics_workbook(Path("metrics.xlsx"), TZ)

    assert calls == [
        (Path("metrics.xlsx"), True, False),
        (Path("metrics.xlsx"), True, True),
    ]
    assert all(workbook.closed for workbook in workbooks)


def test_workbook_closes_when_row_parsing_raises(monkeypatch):
    workbook = FakeWorkbook(
        [HEADERS, replace(VALID_ROW, "曝光", "not-a-count")]
    )
    monkeypatch.setattr(
        "metrics_collector.workbook.load_workbook",
        lambda *args, **kwargs: workbook,
    )

    with pytest.raises(WorkbookFormatError, match=r"row 2.*曝光"):
        parse_metrics_workbook("ignored.xlsx", TZ)

    assert workbook.closed is True


def test_formula_in_required_data_column_is_rejected(tmp_path):
    path = build_workbook(
        tmp_path,
        [replace(VALID_ROW, "曝光", "=1000+191")],
    )

    with pytest.raises(
        WorkbookFormatError,
        match=r"row 2.*曝光.*formula",
    ):
        parse_metrics_workbook(path, TZ)


def corrupt_workbook_xml(destination, source):
    with ZipFile(source) as source_zip, ZipFile(destination, "w") as output_zip:
        for entry in source_zip.infolist():
            content = source_zip.read(entry.filename)
            if entry.filename == "xl/workbook.xml":
                content = b"<workbook><broken>"
            output_zip.writestr(entry, content)


def corrupt_numeric_cell_xml(destination, source):
    with ZipFile(source) as source_zip, ZipFile(destination, "w") as output_zip:
        for entry in source_zip.infolist():
            content = source_zip.read(entry.filename)
            if entry.filename == "xl/worksheets/sheet1.xml":
                valid_value = b"<v>1191</v>"
                assert valid_value in content
                content = content.replace(
                    valid_value,
                    b"<v>not-a-number</v>",
                    1,
                )
            output_zip.writestr(entry, content)


def test_malformed_numeric_cell_raises_contextual_format_error(tmp_path):
    source = build_workbook(tmp_path, [VALID_ROW], filename="source.xlsx")
    path = tmp_path / "malformed-cell.xlsx"
    corrupt_numeric_cell_xml(path, source)

    with pytest.raises(WorkbookFormatError) as exc_info:
        parse_metrics_workbook(path, TZ)

    message = str(exc_info.value)
    assert str(path) in message
    assert "could not read workbook" in message


@pytest.mark.parametrize(
    ("filename", "make_invalid"),
    [
        ("missing.xlsx", lambda path, valid: None),
        ("random.xlsx", lambda path, valid: path.write_bytes(b"random bytes")),
        (
            "truncated.xlsx",
            lambda path, valid: path.write_bytes(valid.read_bytes()[:100]),
        ),
        (
            "unsupported.xls",
            lambda path, valid: path.write_bytes(valid.read_bytes()),
        ),
        (
            "unsupported-content.xlsx",
            lambda path, valid: ZipFile(path, "w").close(),
        ),
        ("invalid-xml.xlsx", corrupt_workbook_xml),
    ],
)
def test_invalid_workbook_files_raise_contextual_format_error(
    tmp_path,
    filename,
    make_invalid,
):
    valid = build_workbook(tmp_path, [VALID_ROW], filename="source.xlsx")
    path = tmp_path / filename
    make_invalid(path, valid)

    with pytest.raises(WorkbookFormatError) as exc_info:
        parse_metrics_workbook(path, TZ)

    message = str(exc_info.value)
    assert str(path) in message
    assert "workbook" in message
